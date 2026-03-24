# -*- coding: utf-8 -*-
r"""
╔══════════════════════════════════════════════════════════════════════╗
║         UNIFICADOR DE PDFs POR CONTRATO  v4.0                       ║
║                                                                      ║
║  Estructura esperada:                                                ║
║    daniel\                    <- RAIZ que se le pasa al script       ║
║      caratulas\               <- PDFs de caratulas fijas             ║
║      informe 1\               <- cada subcarpeta = un informe        ║
║        0001-informe tecnico.pdf                                      ║
║        0001-autorizaciones.pdf                                       ║
║        datos.xlsx             <- se elimina automaticamente          ║
║      informe 2\                                                      ║
║        ...                                                           ║
║                                                                      ║
║  Uso:                                                                ║
║    python unificador_pdf.py C:\daniel                                ║
║                                                                      ║
║  Instalacion:                                                        ║
║    pip install pypdf openpyxl                                        ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import sys
import logging
from pathlib import Path
from datetime import datetime

from pypdf import PdfReader, PdfWriter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════
#  CONFIGURACION DE POSICIONES
#  - keywords   : palabras que deben aparecer en el nombre del archivo
#  - requerido  : True = su ausencia marca el informe como Incompleto
#  - depende_de : None = siempre aplica
#                 "visita"   = solo si hay visita de seguimiento
#                 "calidad"  = solo si hay medicion de calidad
# ══════════════════════════════════════════════════════════════════════

POSICIONES = {
    "1":   {"nombre": "Informe Técnico",
             "keywords": ["informe tecnico","informe_tecnico","inf tecnico"],
             "requerido": True,  "depende_de": None},
    "2":   {"nombre": "Carátula Obligaciones Grales y Especificas",
             "keywords": ["caratula obligaciones","car obligaciones"],
             "requerido": True,  "depende_de": None},
    "3":   {"nombre": "Obligaciones Grales y Especificas",
             "keywords": ["obligaciones generales","obligaciones especificas","oblig generales"],
             "requerido": True,  "depende_de": None},
    "4":   {"nombre": "Carátula Resultados Medición de Acuerdos",
             "keywords": ["caratula resultados","car resultados","caratula medicion acuerdos"],
             "requerido": True,  "depende_de": None},
    "5":   {"nombre": "Acta Indicadores de Gestión",
             "keywords": ["indicadores de gestion","acta indicadores gestion"],
             "requerido": True,  "depende_de": None},
    "5a":  {"nombre": "Acta Indicadores de Calidad",
             "keywords": ["indicadores de calidad","acta indicadores calidad"],
             "requerido": False, "depende_de": "calidad"},
    "5b":  {"nombre": "Carátula Soportes de Medición",
             "keywords": ["caratula soportes de medicion","caratula soportes medicion","car soportes medicion"],
             "requerido": False, "depende_de": "calidad"},
    "5c":  {"nombre": "Soportes de Medición",
             "keywords": ["soportes de medicion","soportes medicion"],
             "requerido": False, "depende_de": "calidad"},
    "6":   {"nombre": "Carátula Relación de Pagos",
             "keywords": ["caratula relacion de pagos","caratula relacion pagos","car relacion pagos"],
             "requerido": True,  "depende_de": None},
    "7":   {"nombre": "Relación de Pagos",
             "keywords": ["relacion de pagos","relacion pagos"],
             "requerido": True,  "depende_de": None},
    "8":   {"nombre": "Carátula Soporte Pago PILA",
             "keywords": ["caratula soporte pila","car pila","caratula soporte pago pila"],
             "requerido": True,  "depende_de": None},
    "9":   {"nombre": "Soporte Pago PILA",
             "keywords": ["soporte pago pila","soporte pila","pago pila"],
             "requerido": True,  "depende_de": None},
    "10":  {"nombre": "Carátula Autorizaciones",
             "keywords": ["caratula autorizaciones","car autorizaciones"],
             "requerido": True,  "depende_de": None},
    "11":  {"nombre": "Autorizaciones",
             "keywords": ["autorizaciones"],
             "requerido": True,  "depende_de": None},
    "12":  {"nombre": "Carátula Habilitación Proveedor",
             "keywords": ["caratula habilitacion","car habilitacion","habilitacion proveedor"],
             "requerido": True,  "depende_de": None},
    "13":  {"nombre": "REPS",
             "keywords": ["reps"],
             "requerido": True,  "depende_de": None},
    "13a": {"nombre": "Carátula Visita",
             "keywords": ["caratula visita","car visita"],
             "requerido": False, "depende_de": "visita"},
    "13b": {"nombre": "Planilla",
             "keywords": ["planilla"],
             "requerido": False, "depende_de": "visita"},
    "13c": {"nombre": "Soporte de Visita",
             "keywords": ["soporte de visita","soporte visita"],
             "requerido": False, "depende_de": "visita"},
    "14":  {"nombre": "Carátula Comunicación",
             "keywords": ["caratula comunicacion","car comunicacion"],
             "requerido": True,  "depende_de": None},
    "15":  {"nombre": "Comunicaciones",
             "keywords": ["comunicaciones"],
             "requerido": True,  "depende_de": None},
    "15a": {"nombre": "Carátula Otros",
             "keywords": ["caratula otros","car otros"],
             "requerido": False, "depende_de": "visita"},
    "15b": {"nombre": "Pantallazo SARLAFT",
             "keywords": ["sarlaft","pantallazo sarlaft"],
             "requerido": False, "depende_de": "visita"},
}

ORDEN = [
    "1","2","3","4","5","5a","5b","5c",
    "6","7","8","9","10","11","12","13",
    "13a","13b","13c","14","15","15a","15b",
]

# Posiciones que activan cada condicion
DETECTORES = {
    "visita":  {"13a","13b","13c","15a","15b"},
    "calidad": {"5a","5b","5c"},
}

EXCLUIR = {"caratulas","_salida"}


# ══════════════════════════════════════════════════════════════════════
#  UTILIDADES
# ══════════════════════════════════════════════════════════════════════

def norm(t):
    tabla = str.maketrans(
        "áéíóúäëïöüàèìòùÁÉÍÓÚÄËÏÖÜÀÈÌÒÙñÑ",
        "aeiouaeiouaeiouAEIOUAEIOUAEIOUnN",
    )
    return t.lower().translate(tabla)


def identificar_pos(nombre):
    stem = norm(Path(nombre).stem)
    mejor, largo = None, 0
    for pid, info in POSICIONES.items():
        for kw in info["keywords"]:
            kw_n = norm(kw)
            if kw_n in stem and len(kw_n) > largo:
                largo = len(kw_n)
                mejor = pid
    return mejor


def pdfs_de(carpeta):
    if not carpeta.exists():
        return []
    return sorted(p for p in carpeta.iterdir()
                  if p.is_file() and p.suffix.lower() == ".pdf")


def borrar_no_pdf(carpeta):
    borrados = []
    for f in list(carpeta.iterdir()):
        if f.is_file() and f.suffix.lower() != ".pdf":
            try:
                f.unlink()
                borrados.append(f.name)
            except OSError as e:
                logging.warning(f"No se pudo eliminar {f}: {e}")
    return borrados


def fusionar(lista, destino):
    writer = PdfWriter()
    errores = []
    for p in lista:
        try:
            for page in PdfReader(str(p)).pages:
                writer.add_page(page)
        except Exception as e:
            errores.append(f"{p.name}: {e}")
            logging.warning(f"PDF ilegible {p}: {e}")
    if not writer.pages:
        return False, errores
    try:
        with open(destino, "wb") as f:
            writer.write(f)
        return True, errores
    except OSError as e:
        return False, errores + [str(e)]


def detectar_condiciones(pdfs):
    """Retorna set de condiciones activas: 'visita', 'calidad', etc."""
    activas = set()
    for p in pdfs:
        pos = identificar_pos(p.name)
        for cond, triggers in DETECTORES.items():
            if pos in triggers:
                activas.add(cond)
    return activas


# ══════════════════════════════════════════════════════════════════════
#  PROCESAMIENTO DE UN INFORME
# ══════════════════════════════════════════════════════════════════════

def procesar_informe(carpeta_informe, carpeta_caratulas, carpeta_salida):
    nombre = carpeta_informe.name
    res = {
        "informe":      nombre,
        "posiciones":   {
            pid: {
                "nombre":     info["nombre"],
                "archivo":    None,
                "presente":   False,
                "depende_de": info["depende_de"],
                "estado_celda": "?",
            }
            for pid, info in POSICIONES.items()
        },
        "sin_posicion": [],
        "eliminados":   [],
        "errores_pdf":  [],
        "condiciones":  set(),
        "estado":       "",
        "pdf_generado": "",
    }

    # 1. Borrar no-PDFs
    res["eliminados"] = borrar_no_pdf(carpeta_informe)

    # 2. PDFs disponibles
    pdfs = pdfs_de(carpeta_informe)

    # 3. Detectar condiciones activas (visita, calidad, etc.)
    res["condiciones"] = detectar_condiciones(pdfs)

    # 4. Mapear PDFs locales a posiciones
    for pdf in pdfs:
        pos = identificar_pos(pdf.name)
        if pos:
            if not res["posiciones"][pos]["presente"]:
                res["posiciones"][pos]["archivo"]  = pdf
                res["posiciones"][pos]["presente"] = True
            else:
                res["sin_posicion"].append(f"{pdf.name}  [dup pos {pos}]")
        else:
            res["sin_posicion"].append(pdf.name)

    # 5. Completar carátulas faltantes con las compartidas
    for pdf in pdfs_de(carpeta_caratulas):
        pos = identificar_pos(pdf.name)
        if pos and not res["posiciones"][pos]["presente"]:
            res["posiciones"][pos]["archivo"]  = pdf
            res["posiciones"][pos]["presente"] = True

    # 6. Construir lista de merge y marcar estado_celda
    lista_merge, faltantes = [], []
    for pid in ORDEN:
        info   = res["posiciones"][pid]
        dep    = info["depende_de"]
        aplica = (dep is None) or (dep in res["condiciones"])

        if not aplica:
            info["estado_celda"] = "N/A"
            continue

        if info["presente"]:
            info["estado_celda"] = "SI"
            lista_merge.append(info["archivo"])
        else:
            info["estado_celda"] = "NO"
            if POSICIONES[pid]["requerido"] or dep in res["condiciones"]:
                faltantes.append(pid)

    # 7. Fusionar
    if lista_merge:
        destino = carpeta_salida / f"{nombre}.pdf"
        ok, errores = fusionar(lista_merge, destino)
        res["errores_pdf"] = errores
        if ok:
            res["pdf_generado"] = destino.name
            res["estado"] = "Unificado Completo" if not faltantes else "Unificado Incompleto"
        else:
            res["estado"] = "Error al unificar"
    else:
        res["estado"] = "Sin PDFs"

    return res


# ══════════════════════════════════════════════════════════════════════
#  REPORTE EXCEL
# ══════════════════════════════════════════════════════════════════════

AZ="1F4E79"; VE="C6EFCE"; VEF="375623"
AM="FFEB9C"; AMF="7D6608"; RO="FFC7CE"
ROF="9C0006"; GR="EDEDED"; GRF="595959"
AZ2="D9E8F5"  # azul suave para condiciones

def borde():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def cel(ws, r, c, v="", bold=False, sz=9, ct="000000",
        fc=None, ha="center", bd=None):
    cell = ws.cell(r, c, v)
    cell.font      = Font(name="Arial", size=sz, bold=bold, color=ct)
    cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=True)
    if fc: cell.fill   = PatternFill("solid", start_color=fc)
    if bd: cell.border = bd
    return cell


def generar_excel(resultados, nombre_analista, ruta):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    B = borde()

    # Titulo
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Reporte de unificación — {nombre_analista}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color=AZ)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("E1:H1")
    ws["E1"] = f"Generado: {datetime.now().strftime('%d/%m/%Y  %H:%M')}"
    ws["E1"].font      = Font(name="Arial", italic=True, size=9, color=GRF)
    ws["E1"].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[1].height = 22

    FH=3; CI=1; CP0=2
    CE = CP0 + len(ORDEN)
    CV = CE+1; CQ = CV+1; CO = CQ+1   # Estado, Visita, Calidad, Obs

    cel(ws, FH, CI, "INFORME / RESPONSABLE", bold=True, sz=9,
        ct="FFFFFF", fc=AZ, bd=B)

    # Colores de encabezado por grupo
    for i, pid in enumerate(ORDEN):
        dep = POSICIONES[pid]["depende_de"]
        if dep == "calidad":
            bg_hdr = "2E6B3E"   # verde oscuro
        elif dep == "visita":
            bg_hdr = "7D4E00"   # naranja oscuro
        else:
            bg_hdr = AZ
        cel(ws, FH, CP0+i,
            f"Pos {pid}\n{POSICIONES[pid]['nombre']}",
            bold=True, sz=8, ct="FFFFFF", fc=bg_hdr, bd=B)

    for col, lbl in [(CE,"ESTADO"),(CV,"VISITA"),(CQ,"CALIDAD"),(CO,"OBSERVACIONES")]:
        cel(ws, FH, col, lbl, bold=True, sz=9, ct="FFFFFF", fc=AZ, bd=B)

    ws.row_dimensions[FH].height = 52

    # Datos
    for idx, res in enumerate(resultados):
        row = FH + 1 + idx
        alt = "EBF3FB" if idx % 2 else None
        cel(ws, row, CI, res["informe"], sz=9, ha="left", fc=alt, bd=B)

        for i, pid in enumerate(ORDEN):
            ec = res["posiciones"][pid].get("estado_celda","?")
            dep = POSICIONES[pid]["depende_de"]
            if   ec=="SI":  fc,ft,bo = VE, VEF, False
            elif ec=="NO":  fc,ft,bo = AM, AMF, True
            elif ec=="N/A": fc,ft,bo = GR, GRF, False
            else:           fc,ft,bo = None,"000000",False
            cel(ws, row, CP0+i, ec, bold=bo, sz=9, ct=ft, fc=fc, bd=B)

        est = res["estado"]
        if   "Completo" in est and "In" not in est: fe,ft = VE,VEF
        elif "Incompleto" in est:                    fe,ft = AM,AMF
        else:                                         fe,ft = RO,ROF
        cel(ws, row, CE, est, bold=True, sz=9, ct=ft, fc=fe, bd=B)

        conds = res.get("condiciones", set())
        cel(ws, row, CV, "SÍ" if "visita"  in conds else "NO", sz=9, bd=B)
        cel(ws, row, CQ, "SÍ" if "calidad" in conds else "NO", sz=9, bd=B)

        obs = []
        if res.get("sin_posicion"):
            obs.append("Sin posición: " + ", ".join(res["sin_posicion"]))
        if res.get("eliminados"):
            obs.append("Eliminados: " + ", ".join(res["eliminados"]))
        if res.get("errores_pdf"):
            obs.append("Errores: " + "; ".join(res["errores_pdf"]))
        cel(ws, row, CO, " | ".join(obs), sz=8, ha="left", bd=B)

    # Anchos
    ws.column_dimensions[get_column_letter(CI)].width  = 44
    for i in range(len(ORDEN)):
        ws.column_dimensions[get_column_letter(CP0+i)].width = 11
    ws.column_dimensions[get_column_letter(CE)].width  = 22
    ws.column_dimensions[get_column_letter(CV)].width  = 9
    ws.column_dimensions[get_column_letter(CQ)].width  = 10
    ws.column_dimensions[get_column_letter(CO)].width  = 55
    ws.freeze_panes = ws.cell(FH+1, CP0)

    # Resumen
    fl = FH + len(resultados) + 3
    completos   = sum(1 for r in resultados if r["estado"]=="Unificado Completo")
    incompletos = sum(1 for r in resultados if "Incompleto" in r["estado"])
    ws.cell(fl,   CI, "RESUMEN").font = Font(name="Arial", bold=True, size=10)
    ws.cell(fl+1, CI, f"Total informes : {len(resultados)}").font = Font(name="Arial", size=9)
    ws.cell(fl+2, CI, f"Completos      : {completos}").font       = Font(name="Arial", size=9, color=VEF)
    ws.cell(fl+3, CI, f"Incompletos    : {incompletos}").font     = Font(name="Arial", size=9, color=AMF)
    ws.cell(fl+4, CI, f"Con error      : {len(resultados)-completos-incompletos}").font = Font(name="Arial", size=9, color=ROF)

    # Leyenda
    fl2 = fl + 6
    ws.cell(fl2, CI, "LEYENDA:").font = Font(name="Arial", bold=True, size=9)
    for j, (lbl, fc, ft, desc) in enumerate([
        ("SI",  VE,  VEF, "Documento presente y unificado"),
        ("NO",  AM,  AMF, "Documento faltante"),
        ("N/A", GR,  GRF, "No aplica para este informe"),
    ]):
        col = CI + 1 + j*2
        cel(ws, fl2, col, lbl, bold=True, sz=9, ct=ft, fc=fc, bd=B)
        cel(ws, fl2, col+1, desc, sz=9, ha="left")

    # Leyenda encabezados
    fl3 = fl2 + 2
    ws.cell(fl3, CI, "Encabezados:").font = Font(name="Arial", bold=True, size=9)
    for j, (lbl, fc, desc) in enumerate([
        ("Siempre",  AZ,       "Posición obligatoria en todos los informes"),
        ("Calidad",  "2E6B3E", "Solo si hay medición de calidad (5a/5b/5c)"),
        ("Visita",   "7D4E00", "Solo si hubo visita de seguimiento (13a-c/15a-b)"),
    ]):
        col = CI + 1 + j*2
        cel(ws, fl3, col, lbl, bold=True, sz=9, ct="FFFFFF", fc=fc, bd=B)
        cel(ws, fl3, col+1, desc, sz=9, ha="left")

    wb.save(str(ruta))


# ══════════════════════════════════════════════════════════════════════
#  FUNCION PRINCIPAL
# ══════════════════════════════════════════════════════════════════════

def main(ruta_raiz_str):
    raiz = Path(ruta_raiz_str)
    if not raiz.exists():
        print(f"\n[ERROR] La ruta no existe: {ruta_raiz_str}")
        sys.exit(1)

    nombre_analista   = raiz.name
    carpeta_caratulas = raiz / "caratulas"
    if not carpeta_caratulas.exists():
        print("[AVISO] No se encontró la carpeta 'caratulas'. Las carátulas compartidas no se aplicarán.")

    carpeta_salida = raiz / "_SALIDA"
    carpeta_pdfs   = carpeta_salida / "PDFs_unificados"
    carpeta_pdfs.mkdir(parents=True, exist_ok=True)

    log_path = carpeta_salida / f"proceso_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    logging.basicConfig(filename=str(log_path), level=logging.INFO,
                        format="%(asctime)s  %(levelname)-8s  %(message)s",
                        encoding="utf-8")
    logging.info(f"Inicio | raiz: {raiz} | analista: {nombre_analista}")

    informes = sorted(
        d for d in raiz.iterdir()
        if d.is_dir() and d.name.lower() not in EXCLUIR
    )
    if not informes:
        print(f"\n[ERROR] No se encontraron carpetas de informes en:\n  {raiz}")
        sys.exit(1)

    sep = "═"*60
    print(f"\n{sep}")
    print(f"  UNIFICADOR DE PDFs  v4.0  —  Analista: {nombre_analista}")
    print(f"{sep}")
    print(f"  Raiz     : {raiz}")
    print(f"  Informes : {len(informes)}")
    print(f"  Salida   : {carpeta_salida}\n")

    resultados = []
    for carpeta_informe in informes:
        print(f"  • {carpeta_informe.name:<50}", end=" ", flush=True)
        res = procesar_informe(carpeta_informe, carpeta_caratulas, carpeta_pdfs)
        resultados.append(res)

        conds_str = ""
        if "visita"  in res["condiciones"]: conds_str += " [visita]"
        if "calidad" in res["condiciones"]: conds_str += " [calidad]"

        ic = "✔" if res["estado"]=="Unificado Completo" else \
             "⚠" if "Incompleto" in res["estado"] else "✖"
        print(f"{ic}  {res['estado']}{conds_str}")

        for x in res["eliminados"]:   print(f"      [eliminado] {x}")
        for x in res["sin_posicion"]: print(f"      [sin pos.]  {x}")
        for x in res["errores_pdf"]:  print(f"      [error PDF] {x}")

        logging.info(f"'{carpeta_informe.name}' -> {res['estado']} condiciones={res['condiciones']}")

    ruta_excel = carpeta_salida / f"reporte_{nombre_analista}.xlsx"
    generar_excel(resultados, nombre_analista, ruta_excel)

    completos   = sum(1 for r in resultados if r["estado"]=="Unificado Completo")
    incompletos = sum(1 for r in resultados if "Incompleto" in r["estado"])
    errores_n   = len(resultados) - completos - incompletos

    print(f"\n{sep}")
    print(f"  FINALIZADO")
    print(f"  Total    : {len(resultados)} informes")
    print(f"  ✔ Completos   : {completos}")
    print(f"  ⚠ Incompletos : {incompletos}")
    print(f"  ✖ Con error   : {errores_n}")
    print(f"\n  Excel -> {ruta_excel}")
    print(f"  PDFs  -> {carpeta_pdfs}")
    print(f"  Log   -> {log_path.name}")
    print(f"{sep}\n")
    logging.info(f"Fin | total={len(resultados)} completos={completos}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("""
USO:    python unificador_pdf.py <ruta_analista>
EJMP:   python unificador_pdf.py C:\\daniel

ESTRUCTURA ESPERADA:
  C:\\daniel\\
    caratulas\\              <- PDFs de caratulas compartidas
    informe 1\\              <- PDFs directamente aqui
      0001-informe tecnico.pdf
      0001-autorizaciones.pdf
    informe 2\\
      ...

REQUIERE:  pip install pypdf openpyxl
""")
        sys.exit(0)
    main(sys.argv[1])