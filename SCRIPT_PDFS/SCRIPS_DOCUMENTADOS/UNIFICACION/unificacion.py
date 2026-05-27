# -*- coding: utf-8 -*-
r"""
╔══════════════════════════════════════════════════════════════════════╗
║         UNIFICADOR DE PDFs POR CONTRATO  v6.2                       ║
║                                                                      ║
║  Uso:  python unificacion.py "C:\ruta\analista"                      ║
║  Req:  pip install pypdf openpyxl                                    ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import sys, logging
from pathlib import Path
from datetime import datetime
from pypdf import PdfReader, PdfWriter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


POSICIONES = {
    "1":   {"nombre": "Informe Técnico",
             "keywords": ["informe tecnico","inf tecnico"],
             "requerido": True,  "depende_de": None},
    "2":   {"nombre": "Carátula Obligaciones",
             "keywords": ["caratula obligaciones","car obligaciones"],
             "requerido": True,  "depende_de": None},
    "3":   {"nombre": "Obligaciones Generales y Especificas",
             "keywords": ["obligaciones generales","obligaciones especificas",
                          "evaluacion de obligaciones","oblig generales",
                          "evaluacion obligaciones contractuales"],
             "requerido": True,  "depende_de": None},
    "4":   {"nombre": "Carátula Resultados Medición",
             "keywords": ["caratula resultados","car resultados"],
             "requerido": True,  "depende_de": None},
    "5":   {"nombre": "Acta Indicadores de Gestión",
             "keywords": ["indicadores de gestion","indicadores gestion",
                          "acta indicadores gestion","acta indicadores de gestion"],
             "requerido": True,  "depende_de": None},
    # ── 5a: presente si hay calidad (con o sin medición) ─────────────
    "5a":  {"nombre": "Acta Indicadores de Calidad",
             "keywords": ["acta de calidad con medicion",
                          "acta de calidad sin medicion",
                          "indicadores de calidad",
                          "indicadores calidad"],
             "requerido": False, "depende_de": "calidad"},
    # ── 5b/5c: solo cuando el acta es CON medición ───────────────────
    "5b":  {"nombre": "Carátula Soportes de Medición",
             "keywords": ["caratula soportes de medicion","caratula soportes medicion",
                          "car soportes medicion"],
             "requerido": False, "depende_de": "calidad_con_medicion"},
    "5c":  {"nombre": "Soportes de Medición",
             "keywords": ["soportes de medicion","soportes medicion"],
             "requerido": False, "depende_de": "calidad_con_medicion"},
    # ── 5d: Indicadores 441 (opcional, sin dependencia) ──────────────
    "5d":  {"nombre": "Indicadores 441",
             "keywords": ["indicadores 441"],
             "requerido": False, "depende_de": None},
    "6":   {"nombre": "Carátula Relación de Pagos",
             "keywords": ["caratula relacion de pagos","caratula relacion pagos",
                          "car relacion pagos"],
             "requerido": True,  "depende_de": None},
    "7":   {"nombre": "Relación de Pagos",
             "keywords": ["relacion de pagos","relacion pagos","informe de pago"],
             "requerido": True,  "depende_de": None},
    "8":   {"nombre": "Carátula Soporte Pago PILA",
             "keywords": ["caratula soporte pila","car pila","caratula soporte pago pila"],
             "requerido": False, "depende_de": "pila"},
    "9":   {"nombre": "Soporte Pago PILA",
             "keywords": ["soporte pago pila","soporte pila","pago pila"],
             "requerido": False, "depende_de": None},
    "10":  {"nombre": "Carátula Autorizaciones",
             "keywords": ["caratula autorizaciones","car autorizaciones"],
             "requerido": True,  "depende_de": None},
    "11":  {"nombre": "Autorizaciones",
             "keywords": ["autorizaciones"],
             "requerido": True,  "depende_de": None},
    "12":  {"nombre": "Carátula Habilitación Proveedor",
             "keywords": ["caratula habilitacion","car habilitacion","habilitacion proveedor"],
             "requerido": True,  "depende_de": None},
    "12a": {"nombre": "Registro Sanitario",
             "keywords": ["registro sanitario"],
             "requerido": False, "depende_de": None},
    "12b": {"nombre": "Certificado de Almacenamiento",
             "keywords": ["certificado de almacenamiento","certificado almacenamiento"],
             "requerido": False, "depende_de": None},
    "13":  {"nombre": "REPS",
             "keywords": ["reps"],
             "requerido": True,  "depende_de": None},
    # ── SECCIÓN VISITA ────────────────────────────────────────────────
    "13a": {"nombre": "Carátula Visita",
             "keywords": ["caratula visita","car visita"],
             "requerido": False, "depende_de": "visita"},
    # Efectiva
    "13b": {"nombre": "Soporte Visita (Efectiva)",
             "keywords": ["soporte visita","soporte de visita"],
             "requerido": False, "depende_de": "efectiva"},
    "13c": {"nombre": "Plantilla Visita (Efectiva)",
             "keywords": ["plantilla visita","plantilla de visita","planilla visita"],
             "requerido": False, "depende_de": "efectiva"},
    # Escalamiento
    "13d": {"nombre": "Requerimiento (Escalamiento)",
             "keywords": ["requerimiento"],
             "requerido": False, "depende_de": "escalamiento"},
    # Inasistencia
    "13e": {"nombre": "Acta de Inasistencia",
             "keywords": ["acta de inasistencia","acta inasistencia"],
             "requerido": False, "depende_de": "inasistencia"},
    # ── SECCIÓN CONSUMO (después de visita, antes de comunicaciones) ──
    "13f": {"nombre": "Carátula Consumo",
             "keywords": ["caratula consumo","car consumo"],
             "requerido": False, "depende_de": "consumo"},
    "13g": {"nombre": "Acta de Consumo",
             "keywords": ["acta consumo","acta de consumo"],
             "requerido": False, "depende_de": None},
    # ── SECCIÓN COMUNICACIONES ────────────────────────────────────────
    "14":  {"nombre": "Carátula Comunicación",
             "keywords": ["caratula comunicacion","car comunicacion"],
             "requerido": False, "depende_de": None},
    # pos 15 = comunicaciones múltiples → manejado por lógica variable
    # ── SECCIÓN OTROS ────────────────────────────────────────────────
    "16a": {"nombre": "Carátula Otros (sección final)",
             "keywords": [],
             "requerido": False, "depende_de": None},
    # pos 16 = otros múltiples → manejado por lógica variable
    "16b": {"nombre": "Pantallazo SARLAFT",
             "keywords": ["sarlaft","pantallazo sarlaft"],
             "requerido": False, "depende_de": "efectiva"},
    "17":  {"nombre": "RIPS",
             "keywords": ["rips"],
             "requerido": False, "depende_de": None},
    "18":  {"nombre": "Indicadores de Calidad del Prestador",
             "keywords": ["indicadores de calidad del prestador","calidad del prestador"],
             "requerido": False, "depende_de": None},
    "19":  {"nombre": "Eventos Adversos",
             "keywords": ["eventos adversos"],
             "requerido": False, "depende_de": None},
}

ORDEN_FIJAS = [
    "1","2","3","4",
    "5","5a","5d","5b","5c",         # 5d = Indicadores 441 (tras soportes de medición)
    "6","7",
    "8","9",
    "10","11",
    "12","12a","12b",
    "13",
    "13a","13c","13b","13d","13e",   # visita
    "13f","13g",                      # consumo (entre visita y comunicaciones)
    "14",
    # <<< comunicaciones múltiples se insertan aquí >>>
    "16a",
    # <<< otros múltiples se insertan aquí >>>
    "16b","17","18","19",
]

DETECTORES = {
    "visita":               {"13a","13b","13c","13d","13e"},
    "efectiva":             {"13b","13c"},
    "escalamiento":         {"13d"},
    "inasistencia":         {"13e"},
    "calidad":              {"5a","5b","5c","5d"},
    "protesis":             {"12a","12b"},
    "pila":                 {"9"},
    "consumo":              {"13g"},
}

# Keywords que distinguen el tipo de acta de calidad
KW_CALIDAD_CON = "acta de calidad con medicion"
KW_CALIDAD_SIN = "acta de calidad sin medicion"

KW_COMUNICACIONES = ["comunicaciones","comunicacion_","comunicaciones_"]
KW_OTROS          = ["otros1","otros2","otros3","otros4","otros5",
                     "otros_1","otros_2","otros_3","otros_4","otros_5",
                     "otros 1","otros 2","otros 3","otros 4","otros 5"]

EXCLUIR = {"caratulas","_salida"}


# ══════════════════════════════════════════════════════════════════════
#  UTILIDADES
# ══════════════════════════════════════════════════════════════════════

def norm(t):
    """Normaliza: minúsculas, sin tildes, guiones/underscores → espacio."""
    tabla = str.maketrans(
        "áéíóúäëïöüàèìòùÁÉÍÓÚÄËÏÖÜÀÈÌÒÙñÑ",
        "aeiouaeiouaeiouAEIOUAEIOUAEIOUnN")
    t = t.lower().translate(tabla)
    t = t.replace("-", " ").replace("_", " ")
    while "  " in t:
        t = t.replace("  ", " ")
    return t.strip()


def identificar_pos(nombre):
    """Devuelve el ID de posición del archivo. Gana la keyword más larga."""
    stem = norm(Path(nombre).stem)
    mejor, largo = None, 0
    for pid, info in POSICIONES.items():
        for kw in info["keywords"]:
            kw_n = norm(kw)
            if kw_n and kw_n in stem and len(kw_n) > largo:
                largo = len(kw_n)
                mejor = pid
    return mejor


def es_calidad_con_medicion(nombre):
    """True si el archivo es un acta de calidad CON medición."""
    return norm(KW_CALIDAD_CON) in norm(Path(nombre).stem)


def es_calidad_sin_medicion(nombre):
    """True si el archivo es un acta de calidad SIN medición."""
    return norm(KW_CALIDAD_SIN) in norm(Path(nombre).stem)


def es_comunicacion(nombre):
    stem = norm(Path(nombre).stem)
    return any(kw in stem for kw in KW_COMUNICACIONES)

def es_otro(nombre):
    stem = norm(Path(nombre).stem)
    if norm("caratula otros") in stem or norm("car otros") in stem:
        return False
    return "otros" in stem

def pdfs_de(carpeta):
    if not carpeta.exists():
        return []
    return sorted(p for p in carpeta.iterdir()
                  if p.is_file() and p.suffix.lower() == ".pdf")

def borrar_no_pdf(carpeta):
    borrados = []
    for f in list(carpeta.iterdir()):
        if f.is_file() and f.suffix.lower() != ".pdf":
            try:
                f.unlink(); borrados.append(f.name)
            except OSError as e:
                logging.warning(f"No se pudo eliminar {f}: {e}")
    return borrados

def fusionar(lista, destino):
    writer, errores = PdfWriter(), []
    for p in lista:
        try:
            for page in PdfReader(str(p)).pages:
                writer.add_page(page)
        except Exception as e:
            errores.append(f"{p.name}: {e}")
            logging.warning(f"PDF ilegible {p}: {e}")
    if not writer.pages:
        return False, errores
    try:
        with open(destino, "wb") as f:
            writer.write(f)
        return True, errores
    except OSError as e:
        return False, errores + [str(e)]

def detectar_condiciones(pdfs):
    """
    Detecta condiciones activas.

    Lógica de calidad:
      - 'calidad'             → hay algún acta de calidad (con o sin medición)
      - 'calidad_con_medicion'→ el acta se llama "acta de calidad con medicion"
      - 'calidad_sin_medicion'→ el acta se llama "acta de calidad sin medicion"

    Cuando es sin medición, 5b y 5c dependen de 'calidad_con_medicion'
    (que no estará activa), por lo que quedan en N/A automáticamente.
    """
    activas = set()
    for p in pdfs:
        pos = identificar_pos(p.name)

        # Detectar sub-tipo de acta de calidad directamente del nombre
        if pos == "5a":
            activas.add("calidad")          # siempre que haya acta de calidad
            if es_calidad_con_medicion(p.name):
                activas.add("calidad_con_medicion")
            elif es_calidad_sin_medicion(p.name):
                activas.add("calidad_sin_medicion")

        for cond, triggers in DETECTORES.items():
            if pos in triggers:
                activas.add(cond)

    if activas & {"efectiva","escalamiento","inasistencia"}:
        activas.add("visita")
    return activas


# ══════════════════════════════════════════════════════════════════════
#  PROCESAMIENTO
# ══════════════════════════════════════════════════════════════════════

def procesar_informe(carpeta_informe, carpeta_caratulas, carpeta_salida):
    nombre = carpeta_informe.name
    res = {
        "informe":        nombre,
        "posiciones":     {
            pid: {"nombre": info["nombre"], "archivo": None,
                  "presente": False, "depende_de": info["depende_de"],
                  "estado_celda": "?"}
            for pid, info in POSICIONES.items()
        },
        "comunicaciones": [],
        "otros":          [],
        "sin_posicion":   [],
        "eliminados":     [],
        "errores_pdf":    [],
        "condiciones":    set(),
        "estado":         "",
        "pdf_generado":   "",
    }

    # 1. Eliminar no-PDFs
    res["eliminados"] = borrar_no_pdf(carpeta_informe)

    # 2. Listar PDFs
    pdfs = pdfs_de(carpeta_informe)

    # 3. Detectar condiciones (incluye calidad_con_medicion / calidad_sin_medicion)
    res["condiciones"] = detectar_condiciones(pdfs)

    # 4. Mapear cada PDF a su posición / comunicaciones / otros
    for pdf in pdfs:
        pos = identificar_pos(pdf.name)
        if pos:
            if not res["posiciones"][pos]["presente"]:
                res["posiciones"][pos]["archivo"]  = pdf
                res["posiciones"][pos]["presente"] = True
            else:
                res["sin_posicion"].append(f"{pdf.name}  [dup pos {pos}]")
        elif es_comunicacion(pdf.name):
            res["comunicaciones"].append(pdf)
        elif es_otro(pdf.name):
            res["otros"].append(pdf)
        else:
            res["sin_posicion"].append(pdf.name)

    # 5. Completar carátulas faltantes con las compartidas
    caratula_otros_compartida = None
    for pdf in pdfs_de(carpeta_caratulas):
        if any(norm(kw) in norm(pdf.stem) for kw in ["caratula otros","car otros"]):
            caratula_otros_compartida = pdf
        pos = identificar_pos(pdf.name)
        if not pos:
            continue
        if pos == "8" and "pila" not in res["condiciones"]:
            continue
        if pos == "13f" and "consumo" not in res["condiciones"]:
            continue
        if not res["posiciones"][pos]["presente"]:
            res["posiciones"][pos]["archivo"]  = pdf
            res["posiciones"][pos]["presente"] = True

    # 6. Construir orden del merge
    lista_merge, faltantes = [], []

    for pid in ORDEN_FIJAS:
        info   = res["posiciones"][pid]
        dep    = info["depende_de"]
        aplica = (dep is None) or (dep in res["condiciones"])

        # Carátula sección Otros (16a): solo si hay algún doc en esa sección
        if pid == "16a":
            hay_otros = (
                bool(res["otros"])
                or res["posiciones"]["16b"]["presente"]
                or res["posiciones"]["17"]["presente"]
                or res["posiciones"]["18"]["presente"]
                or res["posiciones"]["19"]["presente"]
            )
            if not hay_otros:
                info["estado_celda"] = "N/A"
                continue
            if not info["presente"] and caratula_otros_compartida:
                info["archivo"]  = caratula_otros_compartida
                info["presente"] = True

        if not aplica:
            info["estado_celda"] = "N/A"
            continue

        if info["presente"]:
            info["estado_celda"] = "SI"
            lista_merge.append(info["archivo"])
        else:
            info["estado_celda"] = "NO"
            if POSICIONES[pid]["requerido"] or (dep and dep in res["condiciones"]):
                faltantes.append(pid)

        if pid == "14":
            lista_merge.extend(res["comunicaciones"])
        if pid == "16a":
            lista_merge.extend(res["otros"])

    # 7. Fusionar
    if lista_merge:
        destino = carpeta_salida / f"{nombre}.pdf"
        ok, errores = fusionar(lista_merge, destino)
        res["errores_pdf"] = errores
        if ok:
            res["pdf_generado"] = destino.name
            res["estado"] = "Unificado Completo" if not faltantes else "Unificado Incompleto"
        else:
            res["estado"] = "Error al unificar"
    else:
        res["estado"] = "Sin PDFs"

    return res


# ══════════════════════════════════════════════════════════════════════
#  REPORTE EXCEL
# ══════════════════════════════════════════════════════════════════════

AZ="1F4E79"; VE="C6EFCE"; VEF="375623"
AM="FFEB9C"; AMF="7D6608"; RO="FFC7CE"; ROF="9C0006"
GR="EDEDED"; GRF="595959"
VER_OSC="2E6B3E"; NA_OSC="7D4E00"; CA_OSC="5D4037"
CO_OSC="6A1B4D"   # morado oscuro = consumo

def mk_borde():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def cel(ws, r, c, v="", bold=False, sz=9, ct="000000",
        fc=None, ha="center", bd=None):
    cell = ws.cell(r, c, v)
    cell.font      = Font(name="Arial", size=sz, bold=bold, color=ct)
    cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=True)
    if fc: cell.fill   = PatternFill("solid", start_color=fc)
    if bd: cell.border = bd
    return cell

def color_hdr(pid):
    dep = POSICIONES.get(pid, {}).get("depende_de")
    if pid in ("comunicaciones","otros","14","16a"): return CA_OSC
    if pid in ("13f","13g"):                         return CO_OSC
    if pid in ("4", "5", "5d") or dep in ("calidad","calidad_con_medicion"): return VER_OSC
    if dep in ("visita","efectiva","escalamiento","inasistencia"): return NA_OSC
    return AZ

ORDEN_EXCEL = ORDEN_FIJAS + ["comunicaciones","otros"]

def label_col(pid):
    especiales = {"comunicaciones": "Comunicaciones\n(todas)", "otros": "Otros\n(todos)"}
    if pid in especiales: return especiales[pid]
    return f"Pos {pid}\n{POSICIONES[pid]['nombre']}"

def generar_excel(resultados, nombre_analista, ruta):
    wb = Workbook(); ws = wb.active; ws.title = "Reporte"
    B = mk_borde()

    ws.merge_cells("A1:D1")
    ws["A1"] = f"Reporte de unificación — {nombre_analista}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color=AZ)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("E1:H1")
    ws["E1"] = f"Generado: {datetime.now().strftime('%d/%m/%Y  %H:%M')}"
    ws["E1"].font      = Font(name="Arial", italic=True, size=9, color=GRF)
    ws["E1"].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[1].height = 22

    FH=3; CI=1; CP0=2
    N    = len(ORDEN_EXCEL)
    CE   = CP0 + N
    CV   = CE + 1
    CTV  = CV + 1
    CQ   = CTV + 1
    CTCQ = CQ + 1    # TIPO CALIDAD (con/sin medición)
    CPR  = CTCQ + 1
    CPIL = CPR + 1
    CCON = CPIL + 1
    CO   = CCON + 1  # OBSERVACIONES

    cel(ws, FH, CI, "INFORME / RESPONSABLE", bold=True, sz=9, ct="FFFFFF", fc=AZ, bd=B)
    for i, pid in enumerate(ORDEN_EXCEL):
        cel(ws, FH, CP0+i, label_col(pid), bold=True, sz=8,
            ct="FFFFFF", fc=color_hdr(pid), bd=B)
    for col, lbl in [(CE,"ESTADO"),(CV,"VISITA"),(CTV,"TIPO VISITA"),
                     (CQ,"CALIDAD"),(CTCQ,"TIPO CALIDAD"),
                     (CPR,"PROT/ORT"),(CPIL,"PILA"),
                     (CCON,"CONSUMO"),(CO,"OBSERVACIONES")]:
        cel(ws, FH, col, lbl, bold=True, sz=9, ct="FFFFFF", fc=AZ, bd=B)
    ws.row_dimensions[FH].height = 52

    for idx, res in enumerate(resultados):
        row = FH + 1 + idx
        alt = "EBF3FB" if idx % 2 else None
        cel(ws, row, CI, res["informe"], sz=9, ha="left", fc=alt, bd=B)

        for i, pid in enumerate(ORDEN_EXCEL):
            col = CP0 + i
            if pid == "comunicaciones":
                n = len(res["comunicaciones"])
                cel(ws, row, col, str(n) if n else "—", sz=9,
                    ct=(VEF if n else GRF), fc=(VE if n else GR), bd=B)
                continue
            if pid == "otros":
                n = len(res["otros"])
                cel(ws, row, col, str(n) if n else "—", sz=9,
                    ct=(VEF if n else GRF), fc=(VE if n else GR), bd=B)
                continue
            ec = res["posiciones"][pid].get("estado_celda","?")
            if   ec=="SI":  fc,ft,bo = VE,  VEF, False
            elif ec=="NO":  fc,ft,bo = AM,  AMF, True
            elif ec=="N/A": fc,ft,bo = GR,  GRF, False
            else:           fc,ft,bo = None,"000000",False
            cel(ws, row, col, ec, bold=bo, sz=9, ct=ft, fc=fc, bd=B)

        est = res["estado"]
        fe,ft = ((VE,VEF) if "Completo" in est and "In" not in est else
                 (AM,AMF) if "Incompleto" in est else (RO,ROF))
        cel(ws, row, CE, est, bold=True, sz=9, ct=ft, fc=fe, bd=B)

        conds = res.get("condiciones", set())
        cel(ws, row, CV,   "SÍ" if "visita"  in conds else "NO", sz=9, bd=B)

        tipo_v = ("EFECTIVA"     if "efectiva"     in conds else
                  "ESCALAMIENTO" if "escalamiento" in conds else
                  "INASISTENCIA" if "inasistencia" in conds else "—")
        tv_color = ("375623" if tipo_v=="EFECTIVA" else
                    "7D4E00" if tipo_v=="ESCALAMIENTO" else
                    "9C0006" if tipo_v=="INASISTENCIA" else GRF)
        cel(ws, row, CTV,  tipo_v, bold=(tipo_v!="—"), sz=9, ct=tv_color, bd=B)

        # Calidad: SÍ / NO
        cel(ws, row, CQ,   "SÍ" if "calidad" in conds else "NO", sz=9, bd=B)

        # Tipo calidad: CON MEDICIÓN / SIN MEDICIÓN / —
        if "calidad_con_medicion" in conds:
            tipo_cq, cq_color = "CON MEDICIÓN", VEF
        elif "calidad_sin_medicion" in conds:
            tipo_cq, cq_color = "SIN MEDICIÓN", AMF
        else:
            tipo_cq, cq_color = "—", GRF
        cel(ws, row, CTCQ, tipo_cq, bold=(tipo_cq!="—"), sz=9, ct=cq_color, bd=B)

        cel(ws, row, CPR,  "SÍ" if "protesis" in conds else "NO", sz=9, bd=B)
        cel(ws, row, CPIL, "SÍ" if "pila"     in conds else "NO", sz=9,
            ct=("375623" if "pila" in conds else "9C0006"), bd=B)
        cel(ws, row, CCON, "SÍ" if "consumo"  in conds else "NO", sz=9,
            ct=("375623" if "consumo" in conds else GRF), bd=B)

        obs = []
        if res.get("sin_posicion"):
            obs.append("Sin posición: " + ", ".join(res["sin_posicion"]))
        if res.get("eliminados"):
            obs.append("Eliminados: " + ", ".join(res["eliminados"]))
        if res.get("errores_pdf"):
            obs.append("Errores: " + "; ".join(res["errores_pdf"]))
        cel(ws, row, CO, " | ".join(obs), sz=8, ha="left", bd=B)

    ws.column_dimensions[get_column_letter(CI)].width   = 44
    for i in range(N):
        ws.column_dimensions[get_column_letter(CP0+i)].width = 11
    ws.column_dimensions[get_column_letter(CE)].width   = 22
    ws.column_dimensions[get_column_letter(CV)].width   = 9
    ws.column_dimensions[get_column_letter(CTV)].width  = 15
    ws.column_dimensions[get_column_letter(CQ)].width   = 10
    ws.column_dimensions[get_column_letter(CTCQ)].width = 14
    ws.column_dimensions[get_column_letter(CPR)].width  = 11
    ws.column_dimensions[get_column_letter(CPIL)].width = 8
    ws.column_dimensions[get_column_letter(CCON)].width = 10
    ws.column_dimensions[get_column_letter(CO)].width   = 55
    ws.freeze_panes = ws.cell(FH+1, CP0)

    fl = FH + len(resultados) + 3
    completos   = sum(1 for r in resultados if r["estado"]=="Unificado Completo")
    incompletos = sum(1 for r in resultados if "Incompleto" in r["estado"])
    ws.cell(fl,   CI, "RESUMEN").font = Font(name="Arial", bold=True, size=10)
    ws.cell(fl+1, CI, f"Total     : {len(resultados)}").font = Font(name="Arial", size=9)
    ws.cell(fl+2, CI, f"Completos : {completos}").font       = Font(name="Arial", size=9, color=VEF)
    ws.cell(fl+3, CI, f"Incomp.   : {incompletos}").font     = Font(name="Arial", size=9, color=AMF)
    ws.cell(fl+4, CI, f"Error     : {len(resultados)-completos-incompletos}").font = Font(name="Arial", size=9, color=ROF)

    fl2 = fl + 6
    ws.cell(fl2, CI, "LEYENDA:").font = Font(name="Arial", bold=True, size=9)
    for j, (lbl, fc, ft, desc) in enumerate([
        ("SI",  VE, VEF, "Presente y unificado"),
        ("NO",  AM, AMF, "Faltante"),
        ("N/A", GR, GRF, "No aplica"),
    ]):
        col = CI+1+j*2
        cel(ws, fl2, col, lbl, bold=True, sz=9, ct=ft, fc=fc, bd=B)
        cel(ws, fl2, col+1, desc, sz=9, ha="left")

    fl3 = fl2 + 2
    ws.cell(fl3, CI, "Grupos:").font = Font(name="Arial", bold=True, size=9)
    for j, (lbl, fc, desc) in enumerate([
        ("Siempre",    AZ,      "Obligatorio en todos los informes"),
        ("Calidad",    VER_OSC, "Solo si hay medición de calidad (5a/5b/5c/5d)"),
        ("Visita",     NA_OSC,  "Solo si hubo visita (efectiva/escalamiento/inasistencia)"),
        ("Consumo",    CO_OSC,  "Solo si hay Acta de Consumo (13f carátula + 13g acta)"),
        ("Com/Otros",  CA_OSC,  "Comunicaciones y Otros (archivos múltiples)"),
    ]):
        col = CI+1+j*2
        cel(ws, fl3, col, lbl, bold=True, sz=9, ct="FFFFFF", fc=fc, bd=B)
        cel(ws, fl3, col+1, desc, sz=9, ha="left")

    wb.save(str(ruta))


# ══════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════

def main(ruta_raiz_str):
    raiz = Path(ruta_raiz_str)
    if not raiz.exists():
        print(f"\n[ERROR] La ruta no existe: {ruta_raiz_str}"); sys.exit(1)

    nombre_analista   = raiz.name
    carpeta_caratulas = raiz.parent / "caratulas"
    if not carpeta_caratulas.exists():
        print(f"[AVISO] No se encontró la carpeta 'caratulas' en: {carpeta_caratulas}")

    carpeta_salida = raiz / "_SALIDA"
    carpeta_pdfs   = carpeta_salida / "PDFs_unificados"
    carpeta_pdfs.mkdir(parents=True, exist_ok=True)

    log_path = carpeta_salida / f"proceso_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    logging.basicConfig(filename=str(log_path), level=logging.INFO,
                        format="%(asctime)s  %(levelname)-8s  %(message)s",
                        encoding="utf-8")

    informes = sorted(d for d in raiz.iterdir()
                      if d.is_dir() and d.name.lower() not in EXCLUIR)
    if not informes:
        print(f"\n[ERROR] No se encontraron carpetas de informes en:\n  {raiz}")
        sys.exit(1)

    sep = "═"*64
    print(f"\n{sep}")
    print(f"  UNIFICADOR DE PDFs  v6.2  —  Analista: {nombre_analista}")
    print(f"{sep}")
    print(f"  Raiz     : {raiz}")
    print(f"  Informes : {len(informes)}")
    print(f"  Salida   : {carpeta_salida}\n")

    resultados = []
    for carpeta in informes:
        print(f"  • {carpeta.name:<54}", end=" ", flush=True)
        res = procesar_informe(carpeta, carpeta_caratulas, carpeta_pdfs)
        resultados.append(res)

        conds = res["condiciones"]
        info_str = ""
        if   "efectiva"     in conds: info_str += " [visita:EFECTIVA]"
        elif "escalamiento" in conds: info_str += " [visita:ESCALAMIENTO]"
        elif "inasistencia" in conds: info_str += " [visita:INASISTENCIA]"
        # Calidad: mostrar sub-tipo
        if "calidad_con_medicion" in conds: info_str += " [calidad:CON MEDICIÓN]"
        elif "calidad_sin_medicion" in conds: info_str += " [calidad:SIN MEDICIÓN]"
        elif "calidad" in conds: info_str += " [calidad]"
        if "protesis" in conds: info_str += " [protesis/ortesis]"
        if "pila"     in conds: info_str += " [pila]"
        else:                   info_str += " [sin pila]"
        if "consumo"  in conds: info_str += " [consumo]"
        n_com = len(res["comunicaciones"])
        n_otr = len(res["otros"])
        if n_com: info_str += f" [{n_com} comunic.]"
        if n_otr: info_str += f" [{n_otr} otros]"

        ic = ("✔" if res["estado"]=="Unificado Completo" else
              "⚠" if "Incompleto" in res["estado"] else "✖")
        print(f"{ic}  {res['estado']}{info_str}")
        for x in res["eliminados"]:   print(f"      [eliminado] {x}")
        for x in res["sin_posicion"]: print(f"      [sin pos.]  {x}")
        for x in res["errores_pdf"]:  print(f"      [error PDF] {x}")
        logging.info(f"'{carpeta.name}' -> {res['estado']} conds={conds}")

    ruta_excel = carpeta_salida / f"reporte_{nombre_analista}.xlsx"
    generar_excel(resultados, nombre_analista, ruta_excel)

    completos   = sum(1 for r in resultados if r["estado"]=="Unificado Completo")
    incompletos = sum(1 for r in resultados if "Incompleto" in r["estado"])
    print(f"\n{sep}")
    print(f"  FINALIZADO")
    print(f"  Total         : {len(resultados)} informes")
    print(f"  ✔ Completos   : {completos}")
    print(f"  ⚠ Incompletos : {incompletos}")
    print(f"  ✖ Con error   : {len(resultados)-completos-incompletos}")
    print(f"\n  Excel -> {ruta_excel}")
    print(f"  PDFs  -> {carpeta_pdfs}")
    print(f"  Log   -> {log_path.name}")
    print(f"{sep}\n")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("""
USO:    python unificacion.py "<ruta_analista>"
EJMP:   python unificacion.py "C:\\daniel"

ESTRUCTURA:
  analista\\
    caratulas\\
    informe 1\\
    informe 2\\
    ...

REQUIERE:  pip install pypdf openpyxl
"""); sys.exit(0)
    main(sys.argv[1])